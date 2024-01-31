from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.styles.borders import Border, Side, BORDER_THIN
redFill = PatternFill(start_color='DCDCDC',
                   end_color='DCDCDC',
                   fill_type='solid')
redFill1 = PatternFill(start_color='ADD8E6',
                   end_color='ADD8E6',
                   fill_type='solid')
thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='FFFAFA'),
    right=Side(border_style=BORDER_THIN, color='FFFAFA'),
    top=Side(border_style=BORDER_THIN, color='FFFAFA'),
    bottom=Side(border_style=BORDER_THIN, color='FFFAFA')
)
def geral(lista,quantidade):
    wb = Workbook()
    ws = wb.active
    a=3
    b=1
    qtd=int(quantidade)
    qtd=qtd*3
    quantidade2=str(qtd)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width =10
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width =40
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 40
    ws['A2'].fill = redFill1
    ws['A2'].border = thin_border
    ws['B2'].fill = redFill1
    ws['B2'].border = thin_border
    ws['C2'].fill = redFill1
    ws['C2'].border = thin_border
    ws['D2'].fill = redFill1
    ws['D2'].border = thin_border
    ws['E2'].fill = redFill1
    ws['E2'].border = thin_border
    ws['F2'].fill = redFill1
    ws['F2'].border = thin_border
    ws['G2'].fill = redFill1
    ws['G2'].border = thin_border
    ws['H2'].fill = redFill1
    ws['H2'].border = thin_border
    ws['I2'].fill = redFill1
    ws['I2'].border = thin_border
    ws['J2'].fill = redFill1
    ws['J2'].border = thin_border
    ws['K2'].fill = redFill1
    ws['K2'].border = thin_border
    
    ws.row_dimensions[1].height = 60
    
    caminho='A1:k1'
    ws.merge_cells(caminho)
    ws['A1'].value = 'Lista Sistema de dimensionamento dos motores'
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=1, value="Potência_CV")
    ws.cell(row=2, column=2, value="Corrente")
    ws.cell(row=2, column=3, value="Disjuntor motor")
    ws.cell(row=2, column=4, value="TIPO")
    ws.cell(row=2, column=5, value="Suporte de Conexão")
    ws.cell(row=2, column=6, value="Contatora")
    ws.cell(row=2, column=7, value="Contato Auxiliar Frontal")
    ws.cell(row=2, column=8, value="Conector Saque")
    ws.cell(row=2, column=9, value="Seccionador Fusível")
    ws.cell(row=2, column=10, value="Soft Starter")
    ws.cell(row=2, column=11, value="Contatora de Comando")
    ws.row_dimensions[2].height = 60
    c = ws['C3']
    ws.freeze_panes = c
    
    for i in lista:
        
        #print (i[0],i[1],i[2],i[3],i[4])
        vetor=str(i[1])
        vetor=vetor.split(',')
        vetor=vetor[0]

        vetor=float(vetor)
        ws.row_dimensions[a].height = 80
        ws.cell(row=a, column=b, value=i[1])
        
        ws.cell(row=a, column=b+1, value=i[2])
        ws.cell(row=a, column=b+2, value=quantidade+'x'+i[3])
        ws.cell(row=a, column=b+3, value=i[4])
        ws.cell(row=a, column=b+4, value=quantidade+'x'+i[5])
        ws.cell(row=a, column=b+5, value=quantidade+'x'+i[6])
        ws.cell(row=a, column=b+6, value=quantidade+'x'+i[7])
        ws.cell(row=a, column=b+7, value=quantidade2+'x'+i[8])
        if vetor<=15:
            ws.cell(row=a, column=b+8, value=i[9])
            ws.cell(row=a, column=b+9, value=i[10])
            ws.cell(row=a, column=b+10, value=i[11])
        else:
            ws.cell(row=a, column=b+8, value=quantidade+'x'+i[9])
            ws.cell(row=a, column=b+9, value=quantidade+'x'+i[10])
            ws.cell(row=a, column=b+10, value=quantidade+'x'+i[11])
        
        a+=1
    al = Alignment(horizontal="center", vertical="center",wrapText=True)
   
    strvalor='A'+str(a)
    valor=0
    for row in ws['A1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border

            cell.alignment = al
            
        valor+=1
            
    strvalor='B'+str(a)
    valor=0
    for row in ws['B1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border
            cell.alignment = al
        valor+=1
    strvalor='C'+str(a)
    valor=0
    for row in ws['C1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border
            cell.alignment = al
        valor+=1
    strvalor='D'+str(a)
    valor=0
    for row in ws['D1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border
            cell.alignment = al
        valor+=1
    strvalor='E'+str(a)
    valor=0
    for row in ws['E1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border
            cell.alignment = al
        valor+=1
    strvalor='F'+str(a)
    valor=0
    for row in ws['F1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border
            cell.alignment = al
        valor+=1
    strvalor='G'+str(a)
    valor=0
    for row in ws['G1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border
            cell.alignment = al
        valor+=1
    strvalor='H'+str(a)
    valor=0
    for row in ws['H1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border
            cell.alignment = al
        valor+=1
    strvalor='I'+str(a)
    valor=0
    for row in ws['I1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border
            cell.alignment = al
        valor+=1
    strvalor='J'+str(a)
    valor=0
    for row in ws['J1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border
            cell.alignment = al
        valor+=1
    strvalor='K'+str(a)
    valor=0
    for row in ws['K1':strvalor]:
        for cell in row:
            if valor>2:
                if valor%2==1:
                    cell.fill = redFill
                    cell.border = thin_border
            cell.alignment = al
        valor+=1
    ws.title = "RELACAO"
    wb.save("relacao.xlsx")
   


    return "feito"

